"""
Service layer for Cold Storage app.
Contains business logic for data operations, separated from views.
"""
import csv
import json
from io import BytesIO, StringIO
from typing import Dict, List, Tuple, Any, Optional
from django.core.files.uploadedfile import UploadedFile
from django.db.models import QuerySet
from .models import Category, DataItem


class ImportResult:
    """Container for import operation results."""

    def __init__(self):
        self.imported_count: int = 0
        self.errors: List[str] = []
        self.success: bool = False

    def add_error(self, error: str) -> None:
        """Add an error message to the result."""
        self.errors.append(error)

    def increment_imported(self) -> None:
        """Increment the count of successfully imported items."""
        self.imported_count += 1

    def finalize(self) -> None:
        """Mark the import as successful if any items were imported."""
        self.success = self.imported_count > 0

    def get_error_summary(self, max_errors: int = 5) -> str:
        """Get a summary of errors, limiting to max_errors."""
        if not self.errors:
            return ""

        error_message = f'Encountered {len(self.errors)} errors:\n'
        error_message += '\n'.join(self.errors[:max_errors])

        if len(self.errors) > max_errors:
            error_message += f'\n... and {len(self.errors) - max_errors} more errors.'

        return error_message


class JSONImportService:
    """Service for importing data from JSON files."""

    @staticmethod
    def validate_file(file: UploadedFile) -> Tuple[bool, Optional[str]]:
        """
        Validate uploaded file.
        Returns (is_valid, error_message).
        """
        if not file.name.endswith('.json'):
            return False, 'Please upload a JSON file.'
        return True, None

    @staticmethod
    def parse_json_file(file: UploadedFile) -> Tuple[Optional[List[Dict]], Optional[str]]:
        """
        Parse JSON file content.
        Returns (data, error_message).
        """
        try:
            file_content = file.read().decode('utf-8')
            data = json.loads(file_content)
        except json.JSONDecodeError as e:
            return None, f'Invalid JSON format: {str(e)}'
        except UnicodeDecodeError:
            return None, 'File encoding not supported. Please use UTF-8.'

        if not isinstance(data, list):
            return None, 'JSON file must contain an array of objects.'

        return data, None

    @staticmethod
    def get_or_create_category(category_name: str) -> Category:
        """Get existing category or create a new one."""
        category, created = Category.objects.get_or_create(
            name=category_name,
            defaults={'description': f'Auto-created category for {category_name}'}
        )
        return category

    @staticmethod
    def parse_size_estimate(size_value: Any) -> Optional[float]:
        """Parse and validate size estimate value."""
        if size_value is None:
            return None

        try:
            return float(size_value)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def create_data_item_from_entry(entry: Dict[str, Any], category: Category) -> DataItem:
        """Create a DataItem from a JSON entry."""
        return DataItem.objects.create(
            name=entry.get('name', '').strip(),
            category=category,
            description=entry.get('description', ''),
            examples=entry.get('examples', ''),
            size_estimate_gb=JSONImportService.parse_size_estimate(
                entry.get('size_estimate_gb')
            ),
            tags=entry.get('tags', ''),
            source_url=entry.get('source_url', ''),
            notes=entry.get('notes', ''),
            subcategory=entry.get('subcategory', ''),
            priority=entry.get('priority', 'medium'),
            status=entry.get('status', 'planned'),
        )

    @classmethod
    def import_from_json(cls, file: UploadedFile) -> ImportResult:
        """
        Import data items from a JSON file.
        Returns an ImportResult object with statistics and errors.
        """
        result = ImportResult()

        # Validate file
        is_valid, error = cls.validate_file(file)
        if not is_valid:
            result.add_error(error)
            return result

        # Parse JSON
        data, error = cls.parse_json_file(file)
        if error:
            result.add_error(error)
            return result

        # Process each entry
        for i, entry in enumerate(data):
            try:
                if not isinstance(entry, dict):
                    result.add_error(f'Entry {i+1}: Must be an object')
                    continue

                # Validate required fields
                name = entry.get('name', '').strip()
                if not name:
                    result.add_error(f'Entry {i+1}: Name is required')
                    continue

                # Get or create category
                category_name = entry.get('category', 'Uncategorized')
                category = cls.get_or_create_category(category_name)

                # Create data item
                cls.create_data_item_from_entry(entry, category)
                result.increment_imported()

            except Exception as e:
                result.add_error(f'Entry {i+1}: {str(e)}')

        result.finalize()
        return result


class DataItemService:
    """Service for managing DataItem operations."""

    @staticmethod
    def create_from_form_data(form_data: Dict[str, Any]) -> DataItem:
        """Create a DataItem from form data."""
        category = Category.objects.get(id=form_data['category_id'])

        return DataItem.objects.create(
            name=form_data.get('name', ''),
            category=category,
            size_estimate_gb=form_data.get('size_estimate_gb') or None,
            tags=form_data.get('tags', ''),
            description=form_data.get('description', ''),
            subcategory=form_data.get('subcategory', ''),
            source_url=form_data.get('source_url', ''),
            notes=form_data.get('notes', ''),
            examples=form_data.get('examples', ''),
            priority=form_data.get('priority', 'medium'),
            status=form_data.get('status', 'planned'),
        )

    @staticmethod
    def get_statistics() -> Dict[str, Any]:
        """Get overall statistics about data items."""
        from django.db.models import Sum, Count, Avg

        stats = DataItem.objects.aggregate(
            total_items=Count('id'),
            total_size=Sum('size_estimate_gb'),
            average_size=Avg('size_estimate_gb'),
        )

        # Get breakdown by status
        status_breakdown = DataItem.objects.values('status').annotate(
            count=Count('id')
        ).order_by('-count')

        # Get breakdown by priority
        priority_breakdown = DataItem.objects.values('priority').annotate(
            count=Count('id')
        ).order_by('-count')

        return {
            'total_items': stats['total_items'] or 0,
            'total_size_gb': stats['total_size'] or 0,
            'average_size_gb': stats['average_size'] or 0,
            'status_breakdown': list(status_breakdown),
            'priority_breakdown': list(priority_breakdown),
        }

    @staticmethod
    def get_category_statistics() -> List[Dict[str, Any]]:
        """Get statistics broken down by category."""
        from django.db.models import Sum, Count

        return list(
            Category.objects.annotate(
                item_count=Count('data_items'),
                total_size=Sum('data_items__size_estimate_gb')
            ).values(
                'id', 'name', 'item_count', 'total_size'
            ).order_by('-total_size')
        )


class ExportService:
    """Service for exporting data in various formats."""

    @staticmethod
    def _prepare_item_data(item: DataItem) -> Dict[str, Any]:
        """Prepare a single item's data for export."""
        return {
            'id': item.id,
            'name': item.name,
            'category': item.category.name,
            'category_path': item.category.get_full_path(),
            'subcategory': item.subcategory,
            'description': item.description,
            'examples': item.examples,
            'size_estimate_gb': item.size_estimate_gb,
            'size_display': item.get_size_display(),
            'tags': item.tags,
            'source_url': item.source_url,
            'notes': item.notes,
            'priority': item.priority,
            'priority_display': item.get_priority_display(),
            'status': item.status,
            'status_display': item.get_status_display(),
            'created_at': item.created_at.isoformat() if item.created_at else None,
            'updated_at': item.updated_at.isoformat() if item.updated_at else None,
        }

    @classmethod
    def export_to_json(cls, queryset: QuerySet[DataItem]) -> str:
        """
        Export queryset to JSON format.
        Returns JSON string.
        """
        data = [cls._prepare_item_data(item) for item in queryset]
        return json.dumps(data, indent=2, ensure_ascii=False)

    @classmethod
    def export_to_csv(cls, queryset: QuerySet[DataItem]) -> str:
        """
        Export queryset to CSV format.
        Returns CSV string.
        """
        output = StringIO()

        # Define CSV columns
        fieldnames = [
            'id', 'name', 'category', 'category_path', 'subcategory',
            'description', 'examples', 'size_estimate_gb', 'size_display',
            'tags', 'source_url', 'notes', 'priority', 'priority_display',
            'status', 'status_display', 'created_at', 'updated_at'
        ]

        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for item in queryset:
            writer.writerow(cls._prepare_item_data(item))

        return output.getvalue()

    @classmethod
    def export_to_excel(cls, queryset: QuerySet[DataItem]) -> BytesIO:
        """
        Export queryset to Excel format.
        Returns BytesIO object containing Excel file.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Items"

        # Define headers
        headers = [
            'ID', 'Name', 'Category', 'Category Path', 'Subcategory',
            'Description', 'Examples', 'Size (GB)', 'Size Display',
            'Tags', 'Source URL', 'Notes', 'Priority', 'Priority Display',
            'Status', 'Status Display', 'Created At', 'Updated At'
        ]

        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_num, item in enumerate(queryset, 2):
            data = cls._prepare_item_data(item)
            ws.cell(row=row_num, column=1, value=data['id'])
            ws.cell(row=row_num, column=2, value=data['name'])
            ws.cell(row=row_num, column=3, value=data['category'])
            ws.cell(row=row_num, column=4, value=data['category_path'])
            ws.cell(row=row_num, column=5, value=data['subcategory'])
            ws.cell(row=row_num, column=6, value=data['description'])
            ws.cell(row=row_num, column=7, value=data['examples'])
            ws.cell(row=row_num, column=8, value=data['size_estimate_gb'])
            ws.cell(row=row_num, column=9, value=data['size_display'])
            ws.cell(row=row_num, column=10, value=data['tags'])
            ws.cell(row=row_num, column=11, value=data['source_url'])
            ws.cell(row=row_num, column=12, value=data['notes'])
            ws.cell(row=row_num, column=13, value=data['priority'])
            ws.cell(row=row_num, column=14, value=data['priority_display'])
            ws.cell(row=row_num, column=15, value=data['status'])
            ws.cell(row=row_num, column=16, value=data['status_display'])
            ws.cell(row=row_num, column=17, value=data['created_at'])
            ws.cell(row=row_num, column=18, value=data['updated_at'])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @classmethod
    def export_categories_to_json(cls, queryset: QuerySet[Category]) -> str:
        """Export categories to JSON format."""
        data = []
        for category in queryset:
            data.append({
                'id': category.id,
                'name': category.name,
                'description': category.description,
                'parent_id': category.parent_id,
                'parent_name': category.parent.name if category.parent else None,
                'full_path': category.get_full_path(),
                'created_at': category.created_at.isoformat() if category.created_at else None,
                'updated_at': category.updated_at.isoformat() if category.updated_at else None,
            })
        return json.dumps(data, indent=2, ensure_ascii=False)

    @classmethod
    def export_categories_to_csv(cls, queryset: QuerySet[Category]) -> str:
        """Export categories to CSV format."""
        output = StringIO()

        fieldnames = [
            'id', 'name', 'description', 'parent_id', 'parent_name',
            'full_path', 'created_at', 'updated_at'
        ]

        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for category in queryset:
            writer.writerow({
                'id': category.id,
                'name': category.name,
                'description': category.description,
                'parent_id': category.parent_id,
                'parent_name': category.parent.name if category.parent else '',
                'full_path': category.get_full_path(),
                'created_at': category.created_at.isoformat() if category.created_at else '',
                'updated_at': category.updated_at.isoformat() if category.updated_at else '',
            })

        return output.getvalue()
